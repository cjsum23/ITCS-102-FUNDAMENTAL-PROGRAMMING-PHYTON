import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

def display():
    workbook = op.load_workbook('excelDB.xlsx')
    sheet = workbook.active

    for row in tree.get_children():
        tree.delete(row)

    for row in sheet.iter_rows(min_row=2,values_only=True):
        tree.insert('',tk.END,values=row)

def input_validation():
    first = fname_entry.get()
    middle =  mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if not first or not middle or not last or not birth:
        messagebox.showerror('Error','All Fields Required!!')
        return False

    if not birth.isdigit():
        messagebox.showerror('Error','Birth Year Must be a Number')
        return False

    return True

def saving():
    if not input_validation():
        return
    
    fname = fname_entry.get()
    mname = mname_entry.get()
    lname = lname_entry.get()
    birth = int(birth_enrty.get())

    age = 2026-birth

    workbook = op.load_workbook('exelDB.xlsx')
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id,lname,fname,mname,birth,age])

    messagebox.showinfo('Success','Record added successfully')
    display()


window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update")
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=input_validation)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white")
delete_btn.grid(row=6, column=3)

tree = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for col in ("ID","Last","First","Middle","BirthYear","Age"):
    tree.heading(col, text=col)
tree.grid(row=7, column=0, columnspan=4)

display()
window.mainloop()
